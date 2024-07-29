import streamlit as st
import pymongo
import pandas as pd
from fpdf import FPDF
from datetime import datetime
from openpyxl.workbook import Workbook
import io

# Koneksi ke MongoDB
client = pymongo.MongoClient("mongodb+srv://robotikman2jkt:DdlJaVXGJO4Lo91o@cluster0.knymo2f.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0")
db = client["door_lock"] 
collection_log_pintu = db["data_log_pintu_harian"]

st.header("Data Log Akses Pintu")
# Filter ruangan dan tanggal
with st.expander("Filter Data"):
    selected_room = st.selectbox("Pilih Kelas:", collection_log_pintu.distinct("Ruangan"))
    default_date = datetime(2024, 7, 26).date()
    selected_date = st.date_input("Pilih Tanggal:", value=default_date)

# Ambil data log pintu berdasarkan filter
filtered_logs = collection_log_pintu.find({
    "Ruangan": selected_room,
    "Tanggal": selected_date.strftime("%Y-%m-%d")
})
df_logs = pd.DataFrame(list(filtered_logs))

def print_to_pdf(df, selected_room, selected_date):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 10, txt=f"Data Log Akses Ruangan {selected_room} pada {selected_date}", ln=1, align="C")
    pdf.ln(5)

    col_width = pdf.w / 5 

    for col in df[["No.", "Nama", "Ruangan", "Tanggal", "Waktu Akses"]].columns:
        pdf.cell(col_width, 10, txt=col, border=1, align="C")
    pdf.ln()

    for _, row in df[["No.", "Nama", "Ruangan", "Tanggal", "Waktu Akses"]].iterrows():
        for item in row:
            pdf.cell(col_width, 10, txt=str(item), border=1, align="C")
        pdf.ln()

    # Create a temporary file path instead of a file object
    temp_file_path = "temp_log_akses.pdf"  

    # Write the PDF content to the file path
    pdf.output(temp_file_path, "F")

    # Read the file content and convert it to BytesIO for download
    with open(temp_file_path, "rb") as f:
        pdf_output = io.BytesIO(f.read())
    
    # Return the BytesIO object
    return pdf_output

def print_to_xlsx(df):
    xlsx_output = io.BytesIO()
    df[["No.", "Nama", "Ruangan", "Tanggal", "Waktu Akses"]].to_excel(xlsx_output, index=False)
    xlsx_output.seek(0)
    return xlsx_output

# Tampilkan DataFrame
if not df_logs.empty:
    st.write(f"Data Log Akses Ruangan {selected_room} pada {selected_date}")

    st.markdown("""
    <style>
    .dataframe th, .dataframe td {
        text-align: center;
    }
    </style>
    """, unsafe_allow_html=True)

    df_logs.insert(0, "No.", range(1, len(df_logs) + 1))
    df_logs.rename(columns={"timestamp": "Waktu Akses"}, inplace=True)

    # Tombol cetak PDF dan XLSX
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="Cetak ke PDF",
            data=print_to_pdf(df_logs, selected_room, selected_date),
            file_name=f"log_akses_{selected_room}_{selected_date.strftime('%Y-%m-%d')}.pdf",
            mime="application/pdf",
        )
    with col2:
        st.download_button(
            label="Cetak ke XLSX",
            data=print_to_xlsx(df_logs),
            file_name=f"log_akses_{selected_room}_{selected_date.strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.write(df_logs[["No.", "Nama", "Ruangan", "Tanggal", "Waktu Akses"]].to_html(escape=False, index=False), unsafe_allow_html=True)

else:
    st.write("Tidak ada data log akses untuk ruangan dan tanggal yang dipilih.")
